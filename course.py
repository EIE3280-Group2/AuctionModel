from openpyxl import load_workbook
import numpy as np
import copy

wb = load_workbook("course.xlsx")#
ws = wb.active

#添加专业及对应课程要求
#若添加专业，须在第46行模块里添加对应专业的系数操作
#     CSC3170 CSC3180 CSC4001 CSC4008 CSC4140 EIE3001 EIE3280 MAT3007 CSC4080 CHI2012 CHI2017 FRN1002
CE = ["MR",   "MR",   "ME",   "ME",   "ME",   "ME",   "ME",   "ME",   "ME",   "FE",   "FE",   "FE"   ]
CS = ["MR",   "ME",   "MR",   "ME",   "ME",   "ME",   "ME",   "ME",   "ME",   "FE",   "FE",   "FE"   ]
FS = ["ME",   "ME",   "ME",   "ME",   "ME",   "ME",   "ME",   "ME",   "ME",   "FE",   "FE",   "FE"   ]
OT = ["FE",   "FE",   "FE",   "FE",   "FE",   "FE",   "FE",   "FE",   "FE",   "FE",   "FE",   "FE"   ]



quota =[]
head = []
rows = []
jump=2

for row in ws.iter_rows():
    #读最大人数
    if(jump==2):
        jump-=1
        for i in range(0,len(row)-1):
            quota.append(row[i].value)
        continue
    #读课程名称
    if(jump==1):
        jump-=1
        for i in range(0,len(row)-1):
            head.append(row[i].value)
        continue
    if(row[0].value==None):
        break
    d=[]
    #读选课情况
    for i in range(0,len(row)-1):
        d.append(row[i].value if row[i].value!=None else 0)
    rows.append(list(d))
    

#将MR, ME乘上对应系数
for i in range(len(rows)):
    if(rows[i][1]=="CE"):
        for j in range(2,14):
            if CE[j-2]=="MR":
                rows[i][j]*=1.9
            elif CE[j-2]=="ME":
                rows[i][j]*=1.5
        
    if(rows[i][1]=="CS"):
        for j in range(2,14):
            if CS[j-2]=="MR":
                rows[i][j]*=1.9
            elif CS[j-2]=="ME":
                rows[i][j]*=1.5

    if(rows[i][1]=="DS"):
        for j in range(2,14):
            if FS[j-2]=="MR":
                rows[i][j]*=1.9
            elif FS[j-2]=="ME":
                rows[i][j]*=1.5
new=copy.deepcopy(rows)



#对选课结果进行判断
for j in range(2,14):
    d=[]#每位学生出价情况
    num=0#对应课程容量

    for i in range(len(rows)):
        d.append(rows[i][j])
        if(rows[i][j]>0):
            num=num+1
    #若选课人数小于容量，出价人直接选上
    if(num<=quota[j]):
        for i in range(len(rows)):
            rows[i][j]="Yes" if rows[i][j] >0 else "NO"
    #否则，取价最高的前n个人（n为容量）
    else:
        arr = np.array(d)
        b = arr.argsort()[-quota[j]:][::-1]
        for i in range(len(rows)):
            rows[i][j]= "NO"
        for i in b:
            rows[i][j]="Yes"

for j in range(0,len(rows)):
    num=0
    exceed=0
    d=[]
    for i in rows[j]:
        if i=="Yes":
            num=num+1
    if num>6:
        exceed=num-6
        for i in range(2,14):
            
            d.append(new[j][i])
        arr = np.array(d)
        b = arr.argsort()[-6:][::-1]
        for i in range(2,len(rows[j])):
            rows[j][i]="No"
        for i in b:
            rows[j][i+2]="Yes"
print(quota)
print(head)
for i in range(len(rows)):
    print(rows[i])
